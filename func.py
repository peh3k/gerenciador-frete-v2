import requests
import json
import pandas as pd
import openpyxl
from tkinter import ttk
import logging

class DbLink():
    URL_DB = 'https://frete-calculator-default-rtdb.firebaseio.com/'
    PADRAO_DADOS = [
    [
        "ID",
        "nome",
        "peso inicial",
        "peso final",
        "cep inicial",
        "cep final",
        "prazo",
        "estado",
        "cidade",
        "regiao",
        "valor frete",
        "frete min",
        "tac",
        "gris",
        "advalorem",
        "pedagio",
        "tas",
        "icms",
        "outros"
        ], [
            "ID",
            "codigo interno",
            "descricao",
            "unidade",
            "valor venda",
            "peso",
            "comprimento",
            "largura",
            "altura"
        ]
    
    ]

    PATHS = [
        'Transportadora',
        'Produto'
    ]    




def get_db(table):
    requisicao = requests.get(f'{DbLink().URL_DB}/{table}/.json')

    return requisicao.json()

def get_names_db(table):
    all_data = get_db(table)
    names = [lista['nome'] for key in all_data for lista in all_data[key] if lista is not None]

    return names

def get_excel_rows(excel_file):
    df = pd.read_excel(excel_file)
    df.fillna(value="-", inplace=True)
    rows = df.iloc[0:].values.tolist()
    
    return rows



    
def list_to_dict(list_values, padrao_keys):
    lista_de_dicionarios = [dict(zip(padrao_keys, lista)) for lista in list_values]
        
    return lista_de_dicionarios
 


def dicts_to_lists(dicts):
    result = []
    keys = list(dicts[0].keys())
    result.append(keys)
    for d in dicts:
        values = [d[key] for key in keys]
        result.append(values)

    return result

def criar_tabela(frame, colunas, linhas):
    bg_colors = ["white", "gray"]
    frame_tabela = ttk.Frame(frame)
    frame_tabela.pack(fill='both', expand=True)
    style = ttk.Style()
    style.configure("Custom.Treeview", highlightthickness=0,
                    bd=2, relief="groove")
    tree = ttk.Treeview(frame_tabela, style="Custom.Treeview",
                        columns=colunas, show='headings')
    for col in colunas:
        tree.heading(col, text=col)
    for a in linhas:
        tree.insert('', 'end', values=a)
    for col in colunas:
        tree.column(col, width=130, anchor='center')
    if len(colunas) > 1:
        hsb = ttk.Scrollbar(
            frame_tabela, orient='horizontal', command=tree.xview)
        hsb.pack(side='bottom', fill='x')
        tree.configure(xscrollcommand=hsb.set)
    vsb = ttk.Scrollbar(
        frame_tabela, orient='vertical', command=tree.yview)
    vsb.pack(side='right', fill='y')
    tree.configure(yscrollcommand=vsb.set)
    tree.tag_configure("white", background="white")
    tree.tag_configure("gray", background="gray")
    for i, item in enumerate(tree.get_children()):
        bg_color = bg_colors[i % len(bg_colors)]
        tree.item(item, tags=(bg_color,))
    tree.pack(fill='both', expand=True)

def organize_dict(dict, padrao_keys):
    novo_dicionario = {chave: dict[chave] for chave in padrao_keys}
    
    return novo_dicionario



def get_last_id_db(table):
    
    try:
        all_data = get_db(table)
        
        for key in all_data:
            for lista in all_data[key]:
                if lista is None:
                    all_data[key].remove(lista)
        ids = [lista['ID'] for key in all_data for lista in all_data[key]]
        
        return ids[-1]
    except:
        return 0



def patch_db(table, data):
    requisicao = requests.patch(f'{DbLink().URL_DB}/{table}/.json', data=json.dumps(data))

def delete_db(table_path):
    
    try:
        requisicao = requests.delete(f'{DbLink().URL_DB}/{table_path}/.json')
    except IndexError:
        logging.warning("Item não existe")


def criar_tabela_excel(nomes_colunas, linhas, nome_arquivo):
    # Cria um novo arquivo XLSX
    workbook = openpyxl.Workbook()
    # Seleciona a planilha ativa
    sheet = workbook.active
    # Define as colunas com os nomes da lista nomes_colunas
    for coluna, nome_coluna in enumerate(nomes_colunas):
        # A primeira coluna em Excel é "A", portanto, adicionamos 1 ao índice da coluna
        sheet.cell(row=1, column=coluna+1, value=nome_coluna)
    # Define as linhas com os valores da lista linhas
    for linha, valores_linha in enumerate(linhas):
        for coluna, valor in enumerate(valores_linha):
            # A primeira linha em Excel é "1", portanto, adicionamos 2 ao índice da linha
            sheet.cell(row=linha+2, column=coluna+1, value=valor)
    # Salva o arquivo com o nome especificado
    workbook.save(nome_arquivo)


def get_dict_from_name(table, name):
    all_data = get_db(table)
    filtro = [all_data[item].remove(lista) for item in all_data for lista in all_data[item] if lista is None]
    
    
    # Nova lista com nenhuma lista None para interferir
    same_name = [lista for key in all_data for lista in all_data[key] if lista['nome'] == name]
    
    return same_name

def remover_nones(lista):
    
    return list(filter(lambda x: x is not None, lista))



        

    



